from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backend.database import SessionLocal
from backend.models.producto import Producto
from backend.schemas.producto_schema import ProductoCreate, ProductoUpdate

from fastapi import UploadFile, File
import shutil
import uuid

router = APIRouter()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# -------- CREAR / LISTAR --------

@router.post("/productos")
def crear_producto(data: ProductoCreate, db: Session = Depends(get_db)):
    producto = Producto(
        nombre=data.nombre,
        categoria=data.categoria,
        descripcion=data.descripcion,
        precio=data.precio,
        estatus=data.estatus,
        imagen=data.imagen
    )

    db.add(producto)
    db.commit()
    db.refresh(producto)

    return {"msg": "Producto creado correctamente", "producto": producto}


@router.get("/productos")
def obtener_productos(db: Session = Depends(get_db)):
    return db.query(Producto).all()


# -------- RUTAS ESPECIALES PRIMERO --------

@router.get("/productos/buscar/")
def buscar_productos(q: str, db: Session = Depends(get_db)):
    return db.query(Producto).filter(
        or_(
            Producto.nombre.ilike(f"%{q}%"),
            Producto.categoria.ilike(f"%{q}%"),
            Producto.descripcion.ilike(f"%{q}%")
        )
    ).all()


@router.get("/productos/reporte/pdf")
def generar_reporte_productos_pdf(db: Session = Depends(get_db)):
    productos = db.query(Producto).all()
    output_path = "reporte_productos.pdf"

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    contenido = []

    contenido.append(Paragraph("Reporte General de Productos", styles["Title"]))
    contenido.append(Spacer(1, 12))

    if not productos:
        contenido.append(Paragraph("No hay productos registrados.", styles["Normal"]))
    else:
        for producto in productos:
            contenido.append(Paragraph(
                f"<b>{producto.id}. {producto.nombre}</b>",
                styles["Heading3"]
            ))
            contenido.append(Paragraph(f"Categoría: {producto.categoria}", styles["Normal"]))
            contenido.append(Paragraph(f"Descripción: {producto.descripcion or 'Sin descripción'}", styles["Normal"]))
            contenido.append(Paragraph(f"Precio: ${producto.precio}", styles["Normal"]))
            contenido.append(Paragraph(f"Estatus: {producto.estatus}", styles["Normal"]))
            contenido.append(Paragraph(f"Imagen: {producto.imagen or 'Sin imagen'}", styles["Normal"]))
            contenido.append(Spacer(1, 10))

    doc.build(contenido)

    return FileResponse(
        output_path,
        media_type="application/pdf",
        filename="reporte_productos.pdf"
    )


@router.get("/productos/reporte/excel")
def generar_reporte_productos_excel(db: Session = Depends(get_db)):
    productos = db.query(Producto).all()
    output_path = "reporte_productos.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    encabezados = ["ID", "Nombre", "Categoría", "Descripción", "Precio", "Estatus", "Imagen"]
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for producto in productos:
        ws.append([
            producto.id,
            producto.nombre,
            producto.categoria,
            producto.descripcion,
            producto.precio,
            producto.estatus,
            producto.imagen
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 4

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="reporte_productos.xlsx"
    )


@router.get("/dashboard/productos")
def dashboard_productos(db: Session = Depends(get_db)):
    total = db.query(Producto).count()
    existencia = db.query(Producto).filter(Producto.estatus == "Existencia").count()
    agotado = db.query(Producto).filter(Producto.estatus == "Agotado").count()

    categorias = {}

    productos = db.query(Producto).all()

    for producto in productos:
        categoria = producto.categoria or "Sin categoría"
        categorias[categoria] = categorias.get(categoria, 0) + 1

    return {
        "total": total,
        "existencia": existencia,
        "agotado": agotado,
        "por_categoria": categorias
    }


@router.post("/productos/upload-imagen")
def subir_imagen(file: UploadFile = File(...)):
    extension = file.filename.split(".")[-1]
    nombre_archivo = f"{uuid.uuid4()}.{extension}"
    ruta_archivo = f"backend/static/productos/{nombre_archivo}"

    with open(ruta_archivo, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    url = f"http://127.0.0.1:8000/static/productos/{nombre_archivo}"

    return {
        "msg": "Imagen subida correctamente",
        "url": url
    }

# -------- RUTAS CON ID AL FINAL --------

@router.get("/productos/{id}")
def obtener_producto(id: int, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    return producto


@router.get("/productos/{id}/pdf")
def generar_pdf_producto(id: int, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    output_path = f"producto_{producto.id}.pdf"

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    contenido = []

    contenido.append(Paragraph("Ficha Técnica del Producto", styles["Title"]))
    contenido.append(Spacer(1, 12))
    contenido.append(Paragraph(f"<b>ID:</b> {producto.id}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Nombre:</b> {producto.nombre}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Categoría:</b> {producto.categoria}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Descripción:</b> {producto.descripcion or 'Sin descripción'}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Precio:</b> ${producto.precio}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Estatus:</b> {producto.estatus}", styles["Normal"]))
    contenido.append(Paragraph(f"<b>Imagen:</b> {producto.imagen or 'Sin imagen'}", styles["Normal"]))

    doc.build(contenido)

    return FileResponse(
        output_path,
        media_type="application/pdf",
        filename=f"producto_{producto.id}.pdf"
    )


@router.put("/productos/{id}")
def actualizar_producto(id: int, data: ProductoUpdate, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    producto.nombre = data.nombre
    producto.categoria = data.categoria
    producto.descripcion = data.descripcion
    producto.precio = data.precio
    producto.estatus = data.estatus
    producto.imagen = data.imagen

    db.commit()
    db.refresh(producto)

    return {"msg": "Producto actualizado correctamente", "producto": producto}


@router.delete("/productos/{id}")
def eliminar_producto(id: int, db: Session = Depends(get_db)):
    producto = db.query(Producto).filter(Producto.id == id).first()

    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    db.delete(producto)
    db.commit()

    return {"msg": "Producto eliminado correctamente"}