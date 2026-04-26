from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pydantic import BaseModel, EmailStr
from typing import Optional

from backend.database import SessionLocal
from backend.models.sugerencia import Sugerencia

router = APIRouter()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


class SugerenciaCreate(BaseModel):
    nombre: str
    correo: EmailStr
    mensaje: str
    captcha: str


class SugerenciaUpdate(BaseModel):
    estatus: str


@router.post("/sugerencias")
def crear_sugerencia(data: SugerenciaCreate, db: Session = Depends(get_db)):
    if data.captcha.strip() != "1234":
        raise HTTPException(status_code=400, detail="Captcha incorrecto")

    sugerencia = Sugerencia(
        nombre=data.nombre,
        correo=data.correo,
        mensaje=data.mensaje,
        estatus="Pendiente"
    )

    db.add(sugerencia)
    db.commit()
    db.refresh(sugerencia)

    return {"msg": "Sugerencia registrada correctamente", "sugerencia": sugerencia}


@router.get("/sugerencias")
def obtener_sugerencias(skip: int = 0, limit: int = 10, db: Session = Depends(get_db)):
    total = db.query(Sugerencia).count()

    sugerencias = (
        db.query(Sugerencia)
        .order_by(Sugerencia.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "total": total,
        "sugerencias": sugerencias
    }


@router.put("/sugerencias/{id}")
def actualizar_sugerencia(id: int, data: SugerenciaUpdate, db: Session = Depends(get_db)):
    sugerencia = db.query(Sugerencia).filter(Sugerencia.id == id).first()

    if not sugerencia:
        raise HTTPException(status_code=404, detail="Sugerencia no encontrada")

    sugerencia.estatus = data.estatus
    db.commit()
    db.refresh(sugerencia)

    return {"msg": "Sugerencia actualizada", "sugerencia": sugerencia}


@router.delete("/sugerencias/{id}")
def eliminar_sugerencia(id: int, db: Session = Depends(get_db)):
    sugerencia = db.query(Sugerencia).filter(Sugerencia.id == id).first()

    if not sugerencia:
        raise HTTPException(status_code=404, detail="Sugerencia no encontrada")

    db.delete(sugerencia)
    db.commit()

    return {"msg": "Sugerencia eliminada"}


@router.get("/sugerencias/reporte/pdf")
def reporte_sugerencias_pdf(db: Session = Depends(get_db)):
    sugerencias = db.query(Sugerencia).order_by(Sugerencia.id.desc()).all()
    output_path = "reporte_sugerencias.pdf"

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    styles = getSampleStyleSheet()
    contenido = []

    contenido.append(Paragraph("Reporte de Sugerencias", styles["Title"]))
    contenido.append(Spacer(1, 12))

    if not sugerencias:
        contenido.append(Paragraph("No hay sugerencias registradas.", styles["Normal"]))
    else:
        for s in sugerencias:
            contenido.append(Paragraph(f"<b>{s.id}. {s.nombre}</b>", styles["Heading3"]))
            contenido.append(Paragraph(f"Correo: {s.correo}", styles["Normal"]))
            contenido.append(Paragraph(f"Mensaje: {s.mensaje}", styles["Normal"]))
            contenido.append(Paragraph(f"Estatus: {s.estatus}", styles["Normal"]))
            contenido.append(Paragraph(f"Fecha: {s.fecha}", styles["Normal"]))
            contenido.append(Spacer(1, 10))

    doc.build(contenido)

    return FileResponse(
        output_path,
        media_type="application/pdf",
        filename="reporte_sugerencias.pdf"
    )


@router.get("/sugerencias/reporte/excel")
def reporte_sugerencias_excel(db: Session = Depends(get_db)):
    sugerencias = db.query(Sugerencia).order_by(Sugerencia.id.desc()).all()
    output_path = "reporte_sugerencias.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerencias"

    encabezados = ["ID", "Nombre", "Correo", "Mensaje", "Estatus", "Fecha"]
    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for s in sugerencias:
        ws.append([s.id, s.nombre, s.correo, s.mensaje, s.estatus, str(s.fecha)])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 60)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(output_path)

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="reporte_sugerencias.xlsx"
    )